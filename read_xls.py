import openpyxl
import entry
import pandas as pd
from openpyxl import load_workbook


def read_budget_dictionary(budget_dict):
    # Reads item dictionary
    dictionary = pd.read_excel(
        "C:\\Users\\f-eng\\OneDrive\\Documents\\Budget\\dictionary.xlsx", engine="openpyxl")
    # Moves entries from dictionary into a item dictionary.
    for row in range(len(dictionary)):
        key = str(dictionary.loc[row, "ITEM"]).lower().strip()
        new_entry = entry.Entry(key, dictionary.loc[row, "CAT"],
                                dictionary.loc[row, "SUBCAT"], dictionary.loc[row, "BREAKDOWN"])

        budget_dict[new_entry.get_item()] = new_entry


def read_remove_dictionary(remove_dict):
    # Reads remove dictionary
    remove_dictionary = pd.read_excel(
        "C:\\Users\\f-eng\\OneDrive\\Documents\\Budget\\remove_dictionary.xlsx", engine="openpyxl")

    # Moves removed items into a string array
    for row in range(len(remove_dictionary)):
        remove_dict.append(remove_dictionary.loc[row, "ITEM"])


# Parse date field and copy values into day, month, and year columns.
def split_date(monthly_budget):
    for row in range(len(monthly_budget)):
        date = str(monthly_budget.loc[row, "DATE"])
        date_split = date.split("/")
        monthly_budget.loc[row, "DAY"] = int(date_split[1])
        monthly_budget.loc[row, "MONTH"] = int(date_split[0])
        monthly_budget.loc[row, "YEAR"] = int(date_split[2])


def filter_by_month(month, monthly_budget):
    # Remove rows that are not in the current month.
    for index, row in monthly_budget.iterrows():
        if row["MONTH"] != month:
            monthly_budget.drop(index, inplace=True)

    monthly_budget.reset_index(drop=True, inplace=True)


def remove_entries(remove_dict, monthly_budget):
    # Remove entries that match remove dictionary
    for index, row in monthly_budget.iterrows():
        for item in range(len(remove_dict)):
            if remove_dict[item] in row["ITEM"].lower():
                monthly_budget.drop(index, inplace=True)

    monthly_budget.reset_index(drop=True, inplace=True)


def populate_entries(budget_dict, monthly_budget, new_entries):
    # Populate cat, subcat, and breakdown columns with values from budget_dict.
    for row in range(len(monthly_budget)):
        found = False
        for key in budget_dict:
            if key in monthly_budget.loc[row, "ITEM"].lower():
                monthly_budget.loc[row, "CAT"] = budget_dict.get(key).get_cat()
                monthly_budget.loc[row, "SUBCAT"] = budget_dict.get(
                    key).get_subcat()
                monthly_budget.loc[row, "BREAKDOWN"] = budget_dict.get(
                    key).get_breakdown()
                found = True
        if not found:
            # If not in dictionary add to new_entries string array
            new_entries.append(monthly_budget.loc[row, "ITEM"].lower().strip())


def write_to_budget_dictionary(new_entries, budget_dict):
    workbook = openpyxl.load_workbook(
        "C:\\Users\\f-eng\\OneDrive\\Documents\\Budget\\dictionary.xlsx")
    writer = pd.ExcelWriter(
        "C:\\Users\\f-eng\\OneDrive\\Documents\\Budget\\dictionary.xlsx", engine="openpyxl")
    writer.book = workbook
    writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    entries = pd.DataFrame(new_entries)
    entries.to_excel(writer, sheet_name="dict", header=None,
                     index=False, startrow=len(budget_dict))
    writer.save()
    writer.close()


def write_to_budgetfile(monthly_budget):
    workbook = openpyxl.load_workbook(
        "C:\\Users\\f-eng\\OneDrive\\Documents\\Budget\\Accounting_2021.xlsx")
    writer = pd.ExcelWriter(
        "C:\\Users\\f-eng\\OneDrive\\Documents\\Budget\\Accounting_2021.xlsx", engine="openpyxl")
    writer.book = workbook
    writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
    monthly_budget.to_excel(writer, sheet_name="Debit", header=None,
                            index=False, startrow=1)
    writer.save()
    writer.close()


def main():
    budget_dict = {}
    remove_dict = []
    new_entries = []
    monthly_budget = pd.read_excel(
        "C:\\Users\\f-eng\\OneDrive\\Documents\\Budget\\download.xlsx", engine="openpyxl")
    read_budget_dictionary(budget_dict)
    read_remove_dictionary(remove_dict)
    split_date(monthly_budget)
    filter_by_month(1, monthly_budget)
    remove_entries(remove_dict, monthly_budget)
    populate_entries(budget_dict, monthly_budget, new_entries)
    print(monthly_budget.to_string())
    # print(len(monthly_budget))
    # print(new_entries)
    write_to_budget_dictionary(new_entries, budget_dict)
    write_to_budgetfile(monthly_budget)


main()
